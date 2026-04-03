#!/usr/bin/env python3
"""
Export evaluation results to Excel
"""
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Strip illegal XML characters that openpyxl rejects
_ILLEGAL_XML_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]'
)

def sanitize(value):
    """Remove characters that are illegal in Excel/XML cells."""
    if isinstance(value, str):
        return _ILLEGAL_XML_RE.sub('', value)
    return value


# Config (same as compare.py)
RESULT_FILES = {
    'GPT-3.5': [
        'results-gpt35-0-75-final.json',
        'results-gpt35-75-75-final.json',
    ],
    'GPT-4o-mini': [
        'results-gpt4omini-0-75-final.json',
        'results-gpt4omini-75-75-final.json',
    ],
    'GPT-4o-mini+Ctx': [
        'results-gpt4o-context-0-75-final.json',
        'results-gpt4o-context-75-75-final.json',
    ],
}
SHORT = {'GPT-3.5': '3.5', 'GPT-4o-mini': '4o-mini', 'GPT-4o-mini+Ctx': '+Ctx'}

THRESHOLDS = [
    {'name': 'Lenient (BERT>=0.5, F1>=0.5)', 'bert': 0.5, 'f1': 0.5},
    {'name': 'Strict (BERT>=0.7, F1>=0.75)', 'bert': 0.7, 'f1': 0.75},
]

SENTINEL_VALUES = {
    'no question', 'no questions', 'no type', 'no types',
    'no method', 'no methods', 'no collection', 'no analysis',
}
EXCLUDED_QUESTIONS = {
    'data_urls',
    'answer_highlighted',       
    'descriptive_stats_used',  
    'inferential_stats_used',  
    'ml_used',               
}

NO_SIBLING_QUESTIONS = {
    'answer_highlighted', 'data_type', 'descriptive_stats_used',
    'inferential_stats_used', 'method_type', 'ml_used',
    'other_analysis_used', 'rq_text', 'threats_reported',
    'answer_hidden', 
}

# Manual validation overrides (same as compare.py)
FORCE_CORRECT = {
    ('R194401', 'hypothesis_statement'): {'GPT-4o-mini+Ctx'},
    ('R254521', 'method_name_custom'):   {'GPT-4o-mini', 'GPT-4o-mini+Ctx'},
    ('R194431', 'method_name_custom'):   {'GPT-4o-mini', 'GPT-4o-mini+Ctx'},
    ('R195857', 'method_name_custom'):   {'GPT-4o-mini+Ctx'},
    ('R1550970', 'method_name_custom'):  {'GPT-4o-mini+Ctx'},
    ('R584056', 'method_name_custom'):   {'GPT-4o-mini+Ctx'},
    ('R228262', 'method_name_custom'):   {'GPT-4o-mini'},
    ('R211121', 'method_name_custom'):   {'GPT-4o-mini'},
    ('R583135', 'ml_algorithms'):        {'GPT-3.5', 'GPT-4o-mini', 'GPT-4o-mini+Ctx'},
    ('R220605', 'ml_algorithms'):        {'GPT-4o-mini', 'GPT-4o-mini+Ctx'},
    ('R199123', 'statistical_tests'):    {'GPT-4o-mini+Ctx'},
}
EXCLUDE_PAIRS = {
    ('R211145', 'subq_text'),
}

OUTPUT_FILE = 'evaluation-results.xlsx'

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
CORRECT_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
INCORRECT_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def is_sentinel(value):
    if value is None:
        return False
    if isinstance(value, list):
        real = [v for v in value if str(v).lower().strip() not in SENTINEL_VALUES and str(v).strip()]
        return len(real) == 0
    return str(value).lower().strip() in SENTINEL_VALUES


def format_evidence(evidence_list):
    """Format evidence as 'Page X: excerpt' joined by newlines."""
    if not evidence_list:
        return ''
    parts = []
    for ev in evidence_list:
        page = ev.get('pageNumber', '?')
        excerpt = ev.get('excerpt', '')
        parts.append(f"Page {page}: {excerpt}")
    return '\n'.join(parts)


def load_full_questions(paths):
    qs = {}
    for path in paths:
        with open(path) as f:
            data = json.load(f)
        for paper in data['results']:
            pid = paper.get('paperId', '')
            for q in paper['questions']:
                if not q.get('success'):
                    continue
                if is_sentinel(q.get('groundTruth')):
                    continue
                if q.get('questionId') in EXCLUDED_QUESTIONS:
                    continue
                m = q.get('metrics', {})
                if not m or not m.get('suggestions'):
                    continue
                s1 = next((s for s in m['suggestions'] if s.get('position') == 1), None)
                if not s1:
                    continue
                key = (pid, q['questionId'])
                qt = q.get('questionType', '')
                if qt == 'repeat_text':
                    qt = 'text'

                # Extract evidence from top-level suggestions (not metrics)
                evidence = ''
                top_suggestions = q.get('suggestions', [])
                if top_suggestions:
                    top_s1 = next((s for s in top_suggestions if s.get('position') == 1), top_suggestions[0])
                    evidence = format_evidence(top_s1.get('evidence', []))

                qs[key] = {
                    'paperId': pid,
                    'questionId': q['questionId'],
                    'questionType': qt,
                    'groundTruth': q.get('groundTruth'),
                    'prediction': s1.get('text', ''),
                    'bertScore': s1.get('bertScore'),
                    'f1Score': s1.get('f1Score'),
                    'accuracy': s1.get('accuracy'),
                    'isCorrect': s1.get('isCorrect', False),
                    'evidence': evidence,
                }
    return qs


def is_correct_at(q, bert_thresh, f1_thresh):
    qt = q['questionType'].lower()
    if qt in ('boolean', 'select', 'single_select', 'text_object', 'url'):
        return q['isCorrect']
    elif qt == 'text':
        bs = q.get('bertScore')
        return bs is not None and bs >= bert_thresh
    elif qt == 'multi_select':
        f1 = q.get('f1Score')
        return f1 is not None and f1 >= f1_thresh
    return False


def score_value(q):
    qt = q['questionType'].lower()
    if qt in ('boolean', 'select', 'single_select', 'text_object', 'url'):
        return 1.0 if q.get('accuracy') else 0.0
    elif qt == 'text':
        return q.get('bertScore')
    elif qt == 'multi_select':
        return q.get('f1Score')
    return None


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def write_summary_sheet(wb, labels, all_data, common):
    ws = wb.active
    ws.title = 'Summary'

    row = 1
    ws.cell(row=row, column=1, value='Evaluation Summary')
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value=f'Common questions: {len(common)}')
    row += 2

    for thresh in THRESHOLDS:
        ws.cell(row=row, column=1, value=thresh['name'])
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        headers = ['Metric'] + [SHORT[l] for l in labels]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        # S1 accuracy
        ws.cell(row=row, column=1, value='S1 Accuracy')
        for ci, label in enumerate(labels, 2):
            correct = sum(1 for k in common if is_correct_at(all_data[label][k], thresh['bert'], thresh['f1']))
            pct = correct / len(common) * 100
            ws.cell(row=row, column=ci, value=f'{pct:.1f}%')
            ws.cell(row=row, column=ci).alignment = Alignment(horizontal='center')
        row += 1

        # Any correct
        ws.cell(row=row, column=1, value='Any Correct (S1∨S2∨S3)')
        row += 2

    auto_width(ws)


def write_by_type_sheet(wb, labels, all_data, common):
    ws = wb.create_sheet('By Question Type')

    row = 1
    for thresh in THRESHOLDS:
        ws.cell(row=row, column=1, value=thresh['name'])
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        headers = ['Type', 'N'] + [f'{SHORT[l]} Correct' for l in labels] + [f'{SHORT[l]} %' for l in labels]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        by_type = {}
        for k in common:
            qt = all_data[labels[0]][k]['questionType']
            by_type.setdefault(qt, []).append(k)

        for qt in sorted(by_type.keys()):
            keys = by_type[qt]
            n = len(keys)
            ws.cell(row=row, column=1, value=qt)
            ws.cell(row=row, column=2, value=n)
            for ci, label in enumerate(labels):
                correct = sum(1 for k in keys if is_correct_at(all_data[label][k], thresh['bert'], thresh['f1']))
                pct = correct / n * 100 if n else 0
                ws.cell(row=row, column=3 + ci, value=correct)
                ws.cell(row=row, column=3 + len(labels) + ci, value=f'{pct:.1f}%')
                ws.cell(row=row, column=3 + len(labels) + ci).alignment = Alignment(horizontal='center')
            row += 1

        # Total
        total = len(common)
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total)
        for ci, label in enumerate(labels):
            correct = sum(1 for k in common if is_correct_at(all_data[label][k], thresh['bert'], thresh['f1']))
            pct = correct / total * 100
            ws.cell(row=row, column=3 + ci, value=correct)
            ws.cell(row=row, column=3 + len(labels) + ci, value=f'{pct:.1f}%')
        row += 2

    auto_width(ws)


def write_by_qid_sheet(wb, labels, all_data, common):
    ws = wb.create_sheet('By Question ID')

    row = 1
    for thresh in THRESHOLDS:
        ws.cell(row=row, column=1, value=thresh['name'])
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        headers = ['Question ID', 'Type', 'N'] + [f'{SHORT[l]} Correct' for l in labels] + [f'{SHORT[l]} %' for l in labels]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        style_header_row(ws, row, len(headers))
        row += 1

        by_qid = {}
        for k in common:
            qid = all_data[labels[0]][k]['questionId']
            by_qid.setdefault(qid, []).append(k)

        for qid in sorted(by_qid.keys()):
            keys = by_qid[qid]
            n = len(keys)
            qt = all_data[labels[0]][keys[0]]['questionType']
            ws.cell(row=row, column=1, value=qid)
            ws.cell(row=row, column=2, value=qt)
            ws.cell(row=row, column=3, value=n)
            for ci, label in enumerate(labels):
                correct = sum(1 for k in keys if is_correct_at(all_data[label][k], thresh['bert'], thresh['f1']))
                pct = correct / n * 100 if n else 0
                ws.cell(row=row, column=4 + ci, value=correct)
                ws.cell(row=row, column=4 + len(labels) + ci, value=f'{pct:.1f}%')
                ws.cell(row=row, column=4 + len(labels) + ci).alignment = Alignment(horizontal='center')
            row += 1

        # Total
        total = len(common)
        ws.cell(row=row, column=1, value='TOTAL')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total)
        for ci, label in enumerate(labels):
            correct = sum(1 for k in common if is_correct_at(all_data[label][k], thresh['bert'], thresh['f1']))
            pct = correct / total * 100
            ws.cell(row=row, column=4 + ci, value=correct)
            ws.cell(row=row, column=4 + len(labels) + ci, value=f'{pct:.1f}%')
        row += 2

    auto_width(ws)


def write_all_predictions_sheet(wb, labels, all_data, common):
    ws = wb.create_sheet('All Predictions')

    headers = ['Paper ID', 'Question ID', 'Type', 'Ground Truth']
    for label in labels:
        s = SHORT[label]
        headers += [f'{s} Prediction', f'{s} Evidence', f'{s} Score', f'{s} Lenient', f'{s} Strict']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    row = 2
    for k in sorted(common):
        pid, qid = k
        ref = all_data[labels[0]][k]
        gt = ref['groundTruth']
        gt_str = ', '.join(str(x) for x in gt) if isinstance(gt, list) else str(gt)

        ws.cell(row=row, column=1, value=pid)
        ws.cell(row=row, column=2, value=qid)
        ws.cell(row=row, column=3, value=ref['questionType'])
        ws.cell(row=row, column=4, value=gt_str)

        col = 5
        for label in labels:
            q = all_data[label][k]
            pred = str(q['prediction'])[:500]
            evidence = q.get('evidence', '')
            sc = score_value(q)
            lenient = is_correct_at(q, 0.5, 0.5)
            strict = is_correct_at(q, 0.7, 0.75)

            ws.cell(row=row, column=col, value=sanitize(pred))
            ev_cell = ws.cell(row=row, column=col + 1, value=sanitize(evidence[:1000]) if evidence else '')
            ev_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row, column=col + 2, value=round(sc, 4) if sc is not None else '')
            ws.cell(row=row, column=col + 2).alignment = Alignment(horizontal='center')

            lenient_cell = ws.cell(row=row, column=col + 3, value='✓' if lenient else '✗')
            lenient_cell.fill = CORRECT_FILL if lenient else INCORRECT_FILL
            lenient_cell.alignment = Alignment(horizontal='center')

            strict_cell = ws.cell(row=row, column=col + 4, value='✓' if strict else '✗')
            strict_cell.fill = CORRECT_FILL if strict else INCORRECT_FILL
            strict_cell.alignment = Alignment(horizontal='center')

            col += 5

        row += 1

    auto_width(ws, max_width=60)
    for label_idx in range(len(labels)):
        pred_col = 5 + label_idx * 5
        ev_col = 6 + label_idx * 5
        ws.column_dimensions[get_column_letter(pred_col)].width = 60
        ws.column_dimensions[get_column_letter(ev_col)].width = 50


def write_deep_analysis_sheet(wb, labels, all_data, common):
    ws = wb.create_sheet('Deep Analysis')

    headers = ['Question ID', 'Type', 'N', 'Sample Type', 'Paper ID', 'Ground Truth']
    for label in labels:
        s = SHORT[label]
        headers += [f'{s} Prediction', f'{s} Score', f'{s} Correct']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    by_qid = {}
    for k in common:
        qid = all_data[labels[0]][k]['questionId']
        by_qid.setdefault(qid, []).append(k)

    row = 2
    for qid in sorted(by_qid.keys()):
        keys = by_qid[qid]
        qt = all_data[labels[0]][keys[0]]['questionType']
        n = len(keys)

        correct_keys = [k for k in keys if is_correct_at(all_data['GPT-4o-mini'][k], 0.7, 0.75)]
        incorrect_keys = [k for k in keys if not is_correct_at(all_data['GPT-4o-mini'][k], 0.7, 0.75)]

        ws.cell(row=row, column=1, value=qid)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=qt)
        ws.cell(row=row, column=3, value=n)
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).fill = SUBHEADER_FILL
        accs = []
        for label in labels:
            c = sum(1 for k in keys if is_correct_at(all_data[label][k], 0.7, 0.75))
            accs.append(f'{SHORT[label]}: {c}/{n} ({c/n*100:.1f}%)')
        ws.cell(row=row, column=4, value='  |  '.join(accs))
        row += 1

        for sample_keys, sample_type in [(correct_keys[:3], '✓ Correct'), (incorrect_keys[:3], '✗ Incorrect')]:
            for k in sample_keys:
                pid = k[0]
                gt = all_data[labels[0]][k]['groundTruth']
                gt_str = ', '.join(str(x) for x in gt) if isinstance(gt, list) else str(gt)

                ws.cell(row=row, column=4, value=sample_type)
                ws.cell(row=row, column=5, value=pid)
                ws.cell(row=row, column=6, value=gt_str[:200])

                col = 7
                for label in labels:
                    q = all_data[label][k]
                    pred = str(q['prediction'])[:200]
                    sc = score_value(q)
                    strict = is_correct_at(q, 0.7, 0.75)

                    ws.cell(row=row, column=col, value=pred)
                    ws.cell(row=row, column=col + 1, value=round(sc, 4) if sc is not None else '')
                    c_cell = ws.cell(row=row, column=col + 2, value='✓' if strict else '✗')
                    c_cell.fill = CORRECT_FILL if strict else INCORRECT_FILL
                    c_cell.alignment = Alignment(horizontal='center')
                    col += 3

                row += 1

        row += 1 

    auto_width(ws, max_width=60)


def main():
    labels = list(RESULT_FILES.keys())
    all_data = {}
    for label, paths in RESULT_FILES.items():
        all_data[label] = load_full_questions(paths)
        print(f"{label}: {len(all_data[label])} questions")

    ctx_label = 'GPT-4o-mini+Ctx'
    mini_label = 'GPT-4o-mini'
    if ctx_label in all_data and mini_label in all_data:
        copied = 0
        for key, q in all_data[mini_label].items():
            if q['questionId'] in NO_SIBLING_QUESTIONS:
                all_data[ctx_label][key] = q
                copied += 1
        print(f"  Copied {copied} no-sibling questions from {mini_label} → {ctx_label}")

    common = set(all_data[labels[0]].keys())
    for label in labels[1:]:
        common &= set(all_data[label].keys())

    for pair in EXCLUDE_PAIRS:
        common.discard(pair)

    for (pid, qid), model_set in FORCE_CORRECT.items():
        key = (pid, qid)
        for label in model_set:
            if label in all_data and key in all_data[label]:
                q = all_data[label][key]
                q['isCorrect'] = True
                q['bertScore'] = 1.0

    print(f"Common: {len(common)}")

    wb = Workbook()

    print("Writing Summary sheet...")
    write_summary_sheet(wb, labels, all_data, common)

    print("Writing By Question Type sheet...")
    write_by_type_sheet(wb, labels, all_data, common)

    print("Writing By Question ID sheet...")
    write_by_qid_sheet(wb, labels, all_data, common)

    print("Writing All Predictions sheet...")
    write_all_predictions_sheet(wb, labels, all_data, common)

    print("Writing Deep Analysis sheet...")
    write_deep_analysis_sheet(wb, labels, all_data, common)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
